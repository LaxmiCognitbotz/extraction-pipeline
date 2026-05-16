"""Convert NCT report results (JSON) → Excel workbook.

Creates a single sheet with one row per extracted scope item and columns matching
the required reporting structure.
"""

from __future__ import annotations


def write_excel(results: list[dict], output_path: str) -> None:
    """Write a list of NCTReport dicts to an Excel file.

    Args:
        results: List of dicts, each with 'meeting_name' and 'rows'.
        output_path: Path to the .xlsx output file.
    """
    try:
        import openpyxl
    except ImportError:
        # Fallback to CSV if openpyxl not installed
        _write_csv_fallback(results, output_path)
        return

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "NCT Extraction"

    # ── Header styling ──
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = [
        "S.No.",
        "Meeting Name",
        "Transmission Scheme",
        "Transmission Scope",
        "MVA",
        "Status",
        "Source",
        "Approval of Elements in which NCT",
        "Tender Issuing Authority",
        "Date of Tender Issuance",
        "Date of Bid Submission",
        "Execution Timeline (Months)",
        "Tentative SCOD",
        "Awarded To",
        "Project Cost (Cr.)",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # ── Data rows ──
    data_font = Font(name="Calibri", size=10)
    data_align = Alignment(vertical="top", wrap_text=True)
    row_num = 2
    serial = 1

    for result in results:
        meeting = result.get("meeting_name", "")
        rows = result.get("rows", [])

        for elem in rows:
            values = [
                serial,
                meeting,
                elem.get("transmission_scheme", ""),
                elem.get("transmission_scope", ""),
                elem.get("mva"),
                elem.get("status", ""),
                elem.get("source", ""),
                elem.get("approval_of_elements_in_which_nct", ""),
                elem.get("tender_issuing_authority", ""),
                _fmt_date(elem.get("date_of_tender_issuance")),
                _fmt_date(elem.get("date_of_bid_submission")),
                elem.get("execution_timeline_months"),
                _fmt_date(elem.get("tentative_scod")),
                elem.get("awarded_to", ""),
                elem.get("project_cost_cr", ""),
            ]

            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.font = data_font
                cell.alignment = data_align
                cell.border = thin_border

            row_num += 1
            serial += 1

    # ── Column widths ──
    col_widths = [8, 22, 45, 70, 10, 12, 10, 26, 24, 18, 18, 16, 16, 28, 18]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # ── Freeze header row ──
    ws.freeze_panes = "A2"

    # ── Auto-filter ──
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}{row_num - 1}"

    wb.save(output_path)
    print(f"[excel] Wrote {serial - 1} rows to {output_path}")


def _write_csv_fallback(results: list[dict], output_path: str) -> None:
    """Fallback: write CSV if openpyxl is not available."""
    import csv

    csv_path = output_path.replace(".xlsx", ".csv")
    headers = [
        "S.No.",
        "Meeting Name",
        "Transmission Scheme",
        "Transmission Scope",
        "MVA",
        "Status",
        "Source",
        "Approval of Elements in which NCT",
        "Tender Issuing Authority",
        "Date of Tender Issuance",
        "Date of Bid Submission",
        "Execution Timeline (Months)",
        "Tentative SCOD",
        "Awarded To",
        "Project Cost (Cr.)",
    ]

    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(headers)

        serial = 1
        for result in results:
            meeting = result.get("meeting_name", "")
            for elem in result.get("rows", []):
                writer.writerow([
                    serial,
                    meeting,
                    elem.get("transmission_scheme", ""),
                    elem.get("transmission_scope", ""),
                    elem.get("mva", ""),
                    elem.get("status", ""),
                    elem.get("source", ""),
                    elem.get("approval_of_elements_in_which_nct", ""),
                    elem.get("tender_issuing_authority", ""),
                    _fmt_date(elem.get("date_of_tender_issuance")),
                    _fmt_date(elem.get("date_of_bid_submission")),
                    elem.get("execution_timeline_months", ""),
                    _fmt_date(elem.get("tentative_scod")),
                    elem.get("awarded_to", ""),
                    elem.get("project_cost_cr", ""),
                ])
                serial += 1

    print(f"[csv] Wrote {serial - 1} rows to {csv_path} (openpyxl not available)")


def _fmt_date(val) -> str:
    if not val:
        return ""
    # report.to_output_dict() uses ISO strings for date fields.
    if isinstance(val, str):
        return val
    try:
        return val.isoformat()
    except Exception:
        return str(val)
