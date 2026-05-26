"""
Excel converter for CTUIL Revocation (24.6) extraction.

Since column names vary across PDFs, the Excel sheet is built dynamically:
  Fixed columns first: Source File | Upto Month | Application ID | LTA ID | Applicant Name
  Dynamic columns: all unique row_data keys collected across all records, in order of first appearance.
  The result is one unified sheet covering all PDFs.
"""

from __future__ import annotations

import logging
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.revocation_extraction.models import RevocationRecord

logger = logging.getLogger(__name__)

# ── Styles ─────────────────────────────────────────────────────────────────────
_HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
_META_FILL    = PatternFill("solid", fgColor="2E5FA3")
_META_FONT    = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
_DATE_FILL    = PatternFill("solid", fgColor="FFF2CC")
_DATE_FONT    = Font(bold=True, color="7B3F00", name="Calibri", size=10)
_LTA_FILL     = PatternFill("solid", fgColor="E2EFDA")   # green tint for LTA column
_LTA_FONT     = Font(bold=True, color="375623", name="Calibri", size=10)
_DATA_FONT    = Font(name="Calibri", size=10)
_ALT_FILL     = PatternFill("solid", fgColor="EAF0FB")

# Fixed columns always written first
_FIXED_COLS = [
    "source_file",
    "upto_month",
    "application_id",
    "lta_id",
    "applicant_name",
]

_FIXED_LABELS = {
    "source_file":     "Source File",
    "upto_month":      "Upto Month",
    "application_id":  "Application ID",
    "lta_id":          "LTA ID",
    "applicant_name":  "Applicant Name",
}

# Compliance due date column gets amber highlight wherever found
_COMPLIANCE_KEYWORDS = {"24.6", "compliance", "due date", "due"}


def _is_compliance_col(key: str) -> bool:
    lower = key.lower()
    return all(kw in lower for kw in ("24.6", "compliance"))


def records_to_excel(
    records: list[RevocationRecord],
    output_path: Path,
) -> Path:
    """
    Write all records to a single Excel sheet.

    Dynamic column discovery: walks all records first to find every unique
    row_data key, then builds a unified column list for the header row.
    Records that don't have a particular column get a blank cell.
    """
    # ── Collect all unique dynamic column keys in order of first appearance ───
    seen: dict[str, None] = {}
    for rec in records:
        for key in rec.row_data:
            seen.setdefault(key, None)
    dynamic_cols = list(seen.keys())

    all_cols = _FIXED_COLS + dynamic_cols
    logger.info("Excel: %d fixed + %d dynamic = %d total columns",
                len(_FIXED_COLS), len(dynamic_cols), len(all_cols))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Revocations 24.6"

    # ── Header row ────────────────────────────────────────────────────────────
    for col_idx, key in enumerate(all_cols, start=1):
        label = _FIXED_LABELS.get(key, key)   # dynamic cols use raw header name
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        if key == "lta_id":
            cell.fill = _LTA_FILL
            cell.font = _LTA_FONT
        elif key in _FIXED_COLS:
            cell.fill = _META_FILL
            cell.font = _META_FONT
        elif _is_compliance_col(key):
            cell.fill = _DATE_FILL
            cell.font = _DATE_FONT
        else:
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT

    ws.row_dimensions[1].height = 42

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, rec in enumerate(records, start=2):
        d = rec.model_dump()
        row_data: dict = d.get("row_data") or {}
        apply_alt = (row_idx % 2 == 0)

        for col_idx, key in enumerate(all_cols, start=1):
            if key in _FIXED_COLS:
                val = d.get(key)
            else:
                val = row_data.get(key)

            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = _DATA_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            if _is_compliance_col(key) and val:
                cell.fill = PatternFill("solid", fgColor="FFFACD")
            elif key == "lta_id" and val:
                cell.fill = PatternFill("solid", fgColor="F0FFF0")
            elif apply_alt:
                cell.fill = _ALT_FILL

    # ── Column widths ─────────────────────────────────────────────────────────
    width_map = {
        "source_file":    28,
        "upto_month":     12,
        "application_id": 18,
        "lta_id":         35,
        "applicant_name": 38,
    }
    for col_idx, key in enumerate(all_cols, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width_map.get(key, 22)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info("Excel saved: %s  (%d rows, %d columns)", output_path, len(records), len(all_cols))
    return output_path
