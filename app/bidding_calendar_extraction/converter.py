"""
Excel converter for CTUIL Bidding Calendar extraction results.

Outputs ONE single sheet containing every scheme record from every PDF.
Column order:
  Source File | Bidding Calendar Date | Region | Sr. No |
  Transmission Scheme | Major Elements | Bidding Agency |
  Bidding Status | Expected SPV Transfer Date
"""

from __future__ import annotations

import logging
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.bidding_calendar_extraction.models import BiddingSchemeRecord

logger = logging.getLogger(__name__)

# ── Styles ─────────────────────────────────────────────────────────────────
_HEADER_FILL   = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
_META_FILL     = PatternFill("solid", fgColor="2E5FA3")
_META_FONT     = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
_DATE_FILL     = PatternFill("solid", fgColor="FFF2CC")
_DATE_FONT     = Font(bold=True, color="7B3F00", name="Calibri", size=10)
_DATA_FONT     = Font(name="Calibri", size=10)
_ALT_FILL      = PatternFill("solid", fgColor="EAF0FB")

# ── Column layout ──────────────────────────────────────────────────────────
_META_COLS = ["source_file", "bidding_calendar_date", "region"]

_DATA_COLS = [
    "transmission_scheme",
    "major_elements",
    "bidding_agency",
    "bidding_status",
]

_DEADLINE_COLS = ["expected_spv_transfer_date"]

ALL_COLUMNS = _META_COLS + _DATA_COLS + _DEADLINE_COLS

_HEADER_LABELS = {
    "source_file":                 "Source File",
    "bidding_calendar_date":       "Bidding Calendar Date",
    "region":                      "Region",
    "transmission_scheme":         "Transmission Scheme",
    "major_elements":              "Major Elements",
    "bidding_agency":              "Bidding Agency",
    "bidding_status":              "Bidding Status",
    "expected_spv_transfer_date":  "Expected SPV Transfer Date",
}


def _auto_size(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    col_widths = {
        "source_file":                30,
        "bidding_calendar_date":      20,
        "region":                     18,
        "serial_no":                   8,
        "transmission_scheme":        55,
        "major_elements":             60,
        "bidding_agency":             14,
        "bidding_status":             60,
        "expected_spv_transfer_date": 22,
    }
    for col_idx, key in enumerate(ALL_COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = col_widths.get(key, 20)


def records_to_excel(
    records: list[BiddingSchemeRecord],
    output_path: Path,
) -> Path:
    """
    Write all BiddingSchemeRecord objects into a single Excel sheet.

    Each major_element gets its OWN row.
    All other columns (source_file, region, transmission_scheme, etc.) repeat
    on every element row so the sheet is fully self-contained.

    Example — a scheme with 3 elements writes 3 rows:
      Row 2: NR | Scheme X | Element 1 | PFCCL | status | 15.05.2026
      Row 3: NR | Scheme X | Element 2 | PFCCL | status | 15.05.2026
      Row 4: NR | Scheme X | Element 3 | PFCCL | status | 15.05.2026
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bidding Calendar"

    # ── Header row ──────────────────────────────────────────────────────────
    for col_idx, key in enumerate(ALL_COLUMNS, start=1):
        label = _HEADER_LABELS.get(key, key)
        cell = ws.cell(row=1, column=col_idx, value=label)
        if key in _META_COLS:
            cell.fill = _META_FILL
            cell.font = _META_FONT
        elif key in _DEADLINE_COLS:
            cell.fill = _DATE_FILL
            cell.font = _DATE_FONT
        else:
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 36

    # ── Data rows — one row per major_element ────────────────────────────────
    row_idx = 2
    total_written = 0
    scheme_group = 0          # increments per scheme for alternating fill

    for rec in records:
        d = rec.model_dump()
        elements: list[str] = d.get("major_elements") or []

        # If a scheme has no elements, still write one row (major_element = null)
        rows_to_write = elements if elements else [None]
        scheme_group += 1
        apply_alt = (scheme_group % 2 == 0)

        for element in rows_to_write:
            for col_idx, key in enumerate(ALL_COLUMNS, start=1):
                if key == "major_elements":
                    val = element        # one element per row
                else:
                    val = d.get(key)

                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = _DATA_FONT
                cell.alignment = Alignment(vertical="top", wrap_text=True)

                if key in _DEADLINE_COLS and val:
                    cell.fill = PatternFill("solid", fgColor="FFFACD")
                elif apply_alt:
                    cell.fill = _ALT_FILL

            row_idx += 1
            total_written += 1

    if total_written == 0:
        ws.cell(row=2, column=1, value="No records extracted.")
        logger.warning("No rows written to Excel.")

    _auto_size(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info(
        "Excel saved: %s  (%d scheme(s), %d element row(s))",
        output_path, len(records), total_written,
    )
    return output_path

