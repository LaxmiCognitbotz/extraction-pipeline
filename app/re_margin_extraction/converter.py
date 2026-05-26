"""
Excel converter for CTUIL Renewable Energy Margin extraction results.
Outputs 3 different beautifully styled sheets in a single Excel workbook.
Supports multi-row merged headers matching the exact PDF columns layout.
Correctly handles nested Pydantic models with human-readable aliases.
"""

from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# =─────────────────────────────────────────────────────────────────────────────
# Styles & Borders
# =─────────────────────────────────────────────────────────────────────────────

_THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

# 1. Non-RE theme (Ocean Breeze - Deep Blue / Ice Blue)
_NON_RE_HEADER_FILL = PatternFill("solid", fgColor="1F3864")  # Deep Navy
_NON_RE_HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=9)
_NON_RE_ALT_FILL = PatternFill("solid", fgColor="F2F5FA")      # Soft Ice Blue
_NON_RE_STATE_FILL = PatternFill("solid", fgColor="D9E1F2")    # Highlight Blue

# 2. Proposed RE theme (Royal Plum - Majestic Purple / Lavender)
_PROPOSED_RE_HEADER_FILL = PatternFill("solid", fgColor="4A2E80")  # Dark Violet
_PROPOSED_RE_HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=9)
_PROPOSED_RE_ALT_FILL = PatternFill("solid", fgColor="F7F4FC")      # Soft Lavender
_PROPOSED_RE_STATE_FILL = PatternFill("solid", fgColor="E2D9F3")    # Highlight Lavender

# 3. RE Substations theme (Forest Mint - Pine Green / Sage)
_RE_HEADER_FILL = PatternFill("solid", fgColor="1E4D2B")       # Dark Pine Green
_RE_HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=9)
_RE_ALT_FILL = PatternFill("solid", fgColor="F4F8F5")           # Soft Mint
_RE_STATE_FILL = PatternFill("solid", fgColor="D0E1D4")         # Highlight Sage

# Common fonts
_DATA_FONT = Font(name="Calibri", size=10)
_BOLD_DATA_FONT = Font(name="Calibri", size=10, bold=True)

# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────
_ILLEGAL_CHARACTERS_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

def _clean_value(val: any) -> any:
    """Sanitize strings by removing null bytes, tabs, multiple spaces, and openpyxl-illegal characters."""
    if isinstance(val, str):
        # 1. Replace "\t6" sequence with en-dash "–"
        val = val.replace("\t6", "–")
        # 2. Replace tabs with space
        val = val.replace("\t", " ")
        # 3. Remove null bytes
        val = val.replace("\u0000", "")
        # 4. Remove openpyxl-illegal control chars
        val = _ILLEGAL_CHARACTERS_RE.sub("", val)
        # 5. Collapse duplicate spaces
        val = re.sub(r" +", " ", val)
        return val.strip()
    return val


def _auto_size_columns(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """Auto-adjust columns to be wide enough for contents."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 11), 60)


def style_merged_header(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    start_row: int,
    start_col: int,
    end_row: int,
    end_col: int,
    fill: PatternFill,
    font: Font,
    text: str,
    alignment: Alignment = None,
) -> None:
    """Writes text to the top-left cell, merges the range, and styles all cells in the range."""
    # Write top-left
    cell = ws.cell(row=start_row, column=start_col, value=text)
    cell.font = font
    cell.fill = fill
    if alignment:
        cell.alignment = alignment
    cell.border = _THIN_BORDER

    # Style all other cells in the range (so borders and fills are consistent)
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            if r == start_row and c == start_col:
                continue
            other = ws.cell(row=r, column=c)
            other.fill = fill
            other.font = font
            if alignment:
                other.alignment = alignment
            other.border = _THIN_BORDER

    # Merge
    if start_row != end_row or start_col != end_col:
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)


def _get_nested_value(d: dict, path: list[str]) -> any:
    """Safely traverse a nested dictionary using a path of string keys."""
    curr = d
    for step in path:
        if isinstance(curr, dict):
            curr = curr.get(step)
        else:
            return None
    return curr


# =─────────────────────────────────────────────────────────────────────────────
# Sheet 1 Writer: Non-RE
# =─────────────────────────────────────────────────────────────────────────────

NON_RE_COLUMNS = [
    (["Source File"], "Source File"),
    (["As On Date"], "As On Date"),
    (["State"], "State"),
    (["Name of station"], "Name of Station"),
    (["Existing / UC/ Planned MVA Capacity"], "MVA Capacity"),
    (["Capacity Allocated/ Under Process (MW)"], "Capacity Allocated / Under Process (MW)"),
    (["Additional Margin on existing / UC system", "220kV level"], "220kV level"),
    (["Additional Margin on existing / UC system", "400kV level"], "400kV level"),
    (["Line Bays required for RE integration", "220kV level"], "220kV level"),
    (["Line Bays required for RE integration", "400kV level"], "400kV level"),
    (["Additional Margin with ICT Augmentation", "220kV level"], "220kV level"),
    (["Additional Margin with ICT Augmentation", "400kV level"], "400kV level"),
    (["Line Bays required for RE integration (ICT Augmentation)", "220kV level"], "220kV level"),
    (["Line Bays required for RE integration (ICT Augmentation)", "400kV level"], "400kV level"),
    (["No. of Trfs required for RE integration"], "No. of Trfs Required"),
    (["Remarks / Total Addl. Margins"], "Remarks / Total Addl. Margins"),
]

def write_non_re_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, records: list[Any]) -> None:
    ws.title = "Non-RE Substations"
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 24
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 1. Row 1 Main Headers & Merges
    style_merged_header(ws, 1, 1, 2, 1, _NON_RE_HEADER_FILL, _NON_RE_HEADER_FONT, "Source File", align_center)
    style_merged_header(ws, 1, 2, 2, 2, _NON_RE_HEADER_FILL, _NON_RE_HEADER_FONT, "As On Date", align_center)
    style_merged_header(ws, 1, 3, 2, 3, _NON_RE_HEADER_FILL, _NON_RE_HEADER_FONT, "State", align_center)
    style_merged_header(ws, 1, 4, 2, 4, _NON_RE_HEADER_FILL, _NON_RE_HEADER_FONT, "Name of Station", align_center)
    style_merged_header(ws, 1, 5, 2, 5, _NON_RE_HEADER_FILL, _NON_RE_HEADER_FONT, "Existing / UC / Planned MVA Capacity", align_center)
    style_merged_header(ws, 1, 6, 2, 6, _NON_RE_HEADER_FILL, _NON_RE_HEADER_FONT, "Capacity Allocated / Under Process (MW)", align_center)
    
    # Merged categories
    style_merged_header(ws, 1, 7, 1, 8, _NON_RE_HEADER_FILL, _NON_RE_HEADER_FONT, "Addl Margin on Existing/UC System (MW)", align_center)
    style_merged_header(ws, 1, 9, 1, 10, _NON_RE_HEADER_FILL, _NON_RE_HEADER_FONT, "Line Bays Req (Existing System)", align_center)
    style_merged_header(ws, 1, 11, 1, 12, _NON_RE_HEADER_FILL, _NON_RE_HEADER_FONT, "Addl Margin with ICT Aug (MW)", align_center)
    style_merged_header(ws, 1, 13, 1, 14, _NON_RE_HEADER_FILL, _NON_RE_HEADER_FONT, "Line Bays Req (ICT Aug)", align_center)
    
    style_merged_header(ws, 1, 15, 2, 15, _NON_RE_HEADER_FILL, _NON_RE_HEADER_FONT, "No. of Trfs Required", align_center)
    style_merged_header(ws, 1, 16, 2, 16, _NON_RE_HEADER_FILL, _NON_RE_HEADER_FONT, "Remarks / Total Addl. Margins", align_center)

    # 2. Row 2 Sub Headers (Sub columns)
    for col_idx, (_, sub_label) in enumerate(NON_RE_COLUMNS, start=1):
        if col_idx in [7, 8, 9, 10, 11, 12, 13, 14]:
            ws.cell(row=2, column=col_idx, value=sub_label)

    # Write Data starting at row 3
    row_idx = 3
    for r_idx, rec in enumerate(records):
        d = rec.model_dump(by_alias=True)
        apply_alt = (r_idx % 2 == 0)
        
        for col_idx, (path, _) in enumerate(NON_RE_COLUMNS, start=1):
            val = _clean_value(_get_nested_value(d, path))
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = _DATA_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = _THIN_BORDER
            
            if apply_alt:
                cell.fill = _NON_RE_ALT_FILL
            
            # Highlight state column
            if path == ["State"] and val:
                cell.fill = _NON_RE_STATE_FILL
                cell.font = _BOLD_DATA_FONT
        row_idx += 1

    if not records:
        ws.cell(row=3, column=1, value="No records extracted.")
    
    _auto_size_columns(ws)
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = ws.dimensions


# =─────────────────────────────────────────────────────────────────────────────
# Sheet 2 Writer: Proposed RE (Older Non-RE) - Spans 3 Rows
# =─────────────────────────────────────────────────────────────────────────────

PROPOSED_RE_COLUMNS = [
    (["Source File"], "Source File"),
    (["As On Date"], "As On Date"),
    (["State"], "State"),
    (["Name of station"], "Name of Station"),
    (["Transformation Capacity (MVA)", "Existing", "765/400kV"], "765/400kV"),
    (["Transformation Capacity (MVA)", "Existing", "400/220kV or 400/132kV"], "400/220kV or 400/132kV"),
    (["Transformation Capacity (MVA)", "Under Implementation", "765/400kV"], "765/400kV"),
    (["Transformation Capacity (MVA)", "Under Implementation", "400/220kV"], "400/220kV"),
    (["Transformation Capacity (MVA)", "Planned", "765/400kV"], "765/400kV"),
    (["Transformation Capacity (MVA)", "Planned", "400/220kV"], "400/220kV"),
    (["Capacity Allocated (MW)"], "Allocated (MW)"),
    (["Additional Margin on existing / UC system", "220kV level"], "220kV level"),
    (["Additional Margin on existing / UC system", "400kV level"], "400kV level"),
    (["Additional Margin with ICT Augmentation", "220kV level"], "220kV level"),
    (["Additional Margin with ICT Augmentation", "400kV level"], "400kV level"),
    (["No. of Trfs required for RE integration"], "No. of Trfs Required"),
    (["Remarks"], "Remarks"),
]

def write_proposed_re_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, records: list[Any]) -> None:
    ws.title = "Proposed RE Substations"
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 22
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 1. Row 1 Main Headers & Merges (Spans Row 1 to 3)
    style_merged_header(ws, 1, 1, 3, 1, _PROPOSED_RE_HEADER_FILL, _PROPOSED_RE_HEADER_FONT, "Source File", align_center)
    style_merged_header(ws, 1, 2, 3, 2, _PROPOSED_RE_HEADER_FILL, _PROPOSED_RE_HEADER_FONT, "As On Date", align_center)
    style_merged_header(ws, 1, 3, 3, 3, _PROPOSED_RE_HEADER_FILL, _PROPOSED_RE_HEADER_FONT, "State", align_center)
    style_merged_header(ws, 1, 4, 3, 4, _PROPOSED_RE_HEADER_FILL, _PROPOSED_RE_HEADER_FONT, "Name of Station", align_center)
    
    # Transformation Capacity Main Merge
    style_merged_header(ws, 1, 5, 1, 10, _PROPOSED_RE_HEADER_FILL, _PROPOSED_RE_HEADER_FONT, "Transformation Capacity (MVA)", align_center)
    
    style_merged_header(ws, 1, 11, 3, 11, _PROPOSED_RE_HEADER_FILL, _PROPOSED_RE_HEADER_FONT, "Allocated (MW)", align_center)
    
    # Margins merges
    style_merged_header(ws, 1, 12, 1, 13, _PROPOSED_RE_HEADER_FILL, _PROPOSED_RE_HEADER_FONT, "Addl Margin on Existing/UC (MW)", align_center)
    style_merged_header(ws, 1, 14, 1, 15, _PROPOSED_RE_HEADER_FILL, _PROPOSED_RE_HEADER_FONT, "Addl Margin with ICT Aug (MW)", align_center)
    
    style_merged_header(ws, 1, 16, 3, 16, _PROPOSED_RE_HEADER_FILL, _PROPOSED_RE_HEADER_FONT, "No. of Trfs Required", align_center)
    style_merged_header(ws, 1, 17, 3, 17, _PROPOSED_RE_HEADER_FILL, _PROPOSED_RE_HEADER_FONT, "Remarks", align_center)

    # 2. Row 2 Category Merges & Headers
    style_merged_header(ws, 2, 5, 2, 6, _PROPOSED_RE_HEADER_FILL, _PROPOSED_RE_HEADER_FONT, "Existing", align_center)
    style_merged_header(ws, 2, 7, 2, 8, _PROPOSED_RE_HEADER_FILL, _PROPOSED_RE_HEADER_FONT, "Under Implementation", align_center)
    style_merged_header(ws, 2, 9, 2, 10, _PROPOSED_RE_HEADER_FILL, _PROPOSED_RE_HEADER_FONT, "Planned", align_center)
    
    style_merged_header(ws, 2, 12, 3, 12, _PROPOSED_RE_HEADER_FILL, _PROPOSED_RE_HEADER_FONT, "220kV level", align_center)
    style_merged_header(ws, 2, 13, 3, 13, _PROPOSED_RE_HEADER_FILL, _PROPOSED_RE_HEADER_FONT, "400kV level", align_center)
    style_merged_header(ws, 2, 14, 3, 14, _PROPOSED_RE_HEADER_FILL, _PROPOSED_RE_HEADER_FONT, "220kV level", align_center)
    style_merged_header(ws, 2, 15, 3, 15, _PROPOSED_RE_HEADER_FILL, _PROPOSED_RE_HEADER_FONT, "400kV level", align_center)

    # 3. Row 3 Capacity Sub-levels
    for col_idx, (_, sub_label) in enumerate(PROPOSED_RE_COLUMNS, start=1):
        if col_idx in [5, 6, 7, 8, 9, 10]:
            ws.cell(row=3, column=col_idx, value=sub_label)

    # Write Data starting at row 4
    row_idx = 4
    for r_idx, rec in enumerate(records):
        d = rec.model_dump(by_alias=True)
        apply_alt = (r_idx % 2 == 0)
        
        for col_idx, (path, _) in enumerate(PROPOSED_RE_COLUMNS, start=1):
            val = _clean_value(_get_nested_value(d, path))
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = _DATA_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = _THIN_BORDER
            
            if apply_alt:
                cell.fill = _PROPOSED_RE_ALT_FILL
            
            # Highlight state column
            if path == ["State"] and val:
                cell.fill = _PROPOSED_RE_STATE_FILL
                cell.font = _BOLD_DATA_FONT
        row_idx += 1

    if not records:
        ws.cell(row=4, column=1, value="No records extracted.")
    
    _auto_size_columns(ws)
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = ws.dimensions


# =─────────────────────────────────────────────────────────────────────────────
# Sheet 3 Writer: RE Substations
# =─────────────────────────────────────────────────────────────────────────────

RE_COLUMNS = [
    (["Source File"], "Source File"),
    (["As On Date"], "As On Date"),
    (["Region"], "Region"),
    (["Category"], "Category"),
    (["Pooling Station"], "Pooling Station"),
    (["State"], "State"),
    (["RE Potential (MW)", "RE Potential [A]"], "RE Potential [A]"),
    (["RE Potential (MW)", "BESS [B]"], "BESS [B]"),
    (["RE Potential (MW)", "S/s Evacuation Capacity [A-B]"], "S/s Evacuation Capacity [A-B]"),
    (["Expected CoD"], "Expected CoD"),
    (["Connectivity Granted / Agreed (MW)", "220kV"], "220kV"),
    (["Connectivity Granted / Agreed (MW)", "400kV"], "400kV"),
    (["Connectivity Granted / Agreed (MW)", "Total"], "Total (MW)"),
    (["Connectivity Under Process (MW)", "220kV"], "220kV"),
    (["Connectivity Under Process (MW)", "400kV"], "400kV"),
    (["Connectivity Under Process (MW)", "Total"], "Total (MW)"),
    (["Margin for Connectivity (MW)", "220kV"], "220kV"),
    (["Margin for Connectivity (MW)", "400kV"], "400kV"),
    (["Margin for Connectivity (MW)", "Total"], "Total (MW)"),
    (["Additional Margin for Connectivity requiring ICT Augmentation / additional Tr. System (MW)", "220kV"], "220kV"),
    (["Additional Margin for Connectivity requiring ICT Augmentation / additional Tr. System (MW)", "400kV"], "400kV"),
    (["Additional Margin for Connectivity requiring ICT Augmentation / additional Tr. System (MW)", "Total"], "Total (MW)"),
    (["Effectiveness of GNA for Capacity mentioned under 'Margin for Connectivity'"], "Effectiveness of GNA"),
]

def write_re_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, records: list[Any]) -> None:
    ws.title = "RE Substations"
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 24
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 1. Row 1 Main Headers & Merges (Spans Row 1 to 2)
    style_merged_header(ws, 1, 1, 2, 1, _RE_HEADER_FILL, _RE_HEADER_FONT, "Source File", align_center)
    style_merged_header(ws, 1, 2, 2, 2, _RE_HEADER_FILL, _RE_HEADER_FONT, "As On Date", align_center)
    style_merged_header(ws, 1, 3, 2, 3, _RE_HEADER_FILL, _RE_HEADER_FONT, "Region", align_center)
    style_merged_header(ws, 1, 4, 2, 4, _RE_HEADER_FILL, _RE_HEADER_FONT, "Category", align_center)
    style_merged_header(ws, 1, 5, 2, 5, _RE_HEADER_FILL, _RE_HEADER_FONT, "Pooling Station", align_center)
    style_merged_header(ws, 1, 6, 2, 6, _RE_HEADER_FILL, _RE_HEADER_FONT, "State", align_center)
    
    # Merged blocks
    style_merged_header(ws, 1, 7, 1, 9, _RE_HEADER_FILL, _RE_HEADER_FONT, "RE Potential (MW)", align_center)
    style_merged_header(ws, 1, 10, 2, 10, _RE_HEADER_FILL, _RE_HEADER_FONT, "Expected CoD of Pooling Station", align_center)
    style_merged_header(ws, 1, 11, 1, 13, _RE_HEADER_FILL, _RE_HEADER_FONT, "Connectivity Granted / Agreed (MW)", align_center)
    style_merged_header(ws, 1, 14, 1, 16, _RE_HEADER_FILL, _RE_HEADER_FONT, "Connectivity Under Process (MW)", align_center)
    style_merged_header(ws, 1, 17, 1, 19, _RE_HEADER_FILL, _RE_HEADER_FONT, "Margin for Connectivity (MW)", align_center)
    style_merged_header(ws, 1, 20, 1, 22, _RE_HEADER_FILL, _RE_HEADER_FONT, "Addl Margin requiring ICT Aug / Tr. System (MW)", align_center)
    
    style_merged_header(ws, 1, 23, 2, 23, _RE_HEADER_FILL, _RE_HEADER_FONT, "Effectiveness of GNA for Capacity mentioned under 'Margin for Connectivity'", align_center)

    # 2. Row 2 Sub Columns
    for col_idx, (_, sub_label) in enumerate(RE_COLUMNS, start=1):
        if col_idx in [7, 8, 9, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22]:
            ws.cell(row=2, column=col_idx, value=sub_label)

    # Write Data starting at row 3
    row_idx = 3
    for r_idx, rec in enumerate(records):
        d = rec.model_dump(by_alias=True)
        apply_alt = (r_idx % 2 == 0)
        
        for col_idx, (path, _) in enumerate(RE_COLUMNS, start=1):
            val = _clean_value(_get_nested_value(d, path))
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = _DATA_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = _THIN_BORDER
            
            if apply_alt:
                cell.fill = _RE_ALT_FILL
            
            # Highlight state, region, and category columns
            if path[0] in ["State", "Region", "Category"] and val:
                cell.fill = _RE_STATE_FILL
                cell.font = _BOLD_DATA_FONT
        row_idx += 1

    if not records:
        ws.cell(row=3, column=1, value="No records extracted.")
    
    _auto_size_columns(ws)
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = ws.dimensions


# =─────────────────────────────────────────────────────────────────────────────
# Main Workbook Exporter
# =─────────────────────────────────────────────────────────────────────────────

def margins_to_excel(
    non_re_records: list[Any],
    proposed_re_records: list[Any],
    re_records: list[Any],
    output_path: Path,
) -> Path:
    """
    Write all Renewable Energy Margin records into 3 beautifully styled sheets in a single Excel file.
    Supports exact multi-row merged header representations from the original PDFs.
    """
    wb = openpyxl.Workbook()
    
    # 1. Non-RE Substations
    ws_non_re = wb.active
    write_non_re_sheet(ws_non_re, non_re_records)
    
    # 2. Proposed RE Substations
    ws_proposed_re = wb.create_sheet()
    write_proposed_re_sheet(ws_proposed_re, proposed_re_records)
    
    # 3. RE Substations
    ws_re = wb.create_sheet()
    write_re_sheet(ws_re, re_records)
    
    # Ensure parent output path exists
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    
    logger.info(
        "Excel exported: %s with 3 sheets (%d Non-RE, %d Proposed RE, %d RE Substation records)",
        output_path, len(non_re_records), len(proposed_re_records), len(re_records)
    )
    return output_path
