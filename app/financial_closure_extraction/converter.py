"""
Excel converter for CTUIL Compliance extraction results.

Outputs ONE single sheet containing every row from every table across
all processed PDF files. Each row includes metadata columns (source_file,
report_period, table_type, table_name) prepended so rows are self-describing.

Column order:
  source_file | report_period | table_type | table_name |
  application_id | name_of_applicant | submission_date | region |
  location_of_project | type_of_project | installed_capacity_mw |
  first_scod_of_generation_project | connectivity_granted_mw | substation |
  date_of_connectivity_intimation_in_principle |
  date_of_connectivity_intimation_final |
  connectivity_gna_start_date_in_principle |
  connectivity_gna_start_date_firm |
  criterion_for_applying | revised_criterion | revised_scod |
  application_status |
  due_date_of_fc | due_date_for_submission_of_land_docs
"""

from __future__ import annotations

import logging
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.financial_closure_extraction.models import FileExtractionResult

logger = logging.getLogger(__name__)

# ── Styles ────────────────────────────────────────────────────────────────────
_HEADER_FILL    = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT    = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
_META_FILL      = PatternFill("solid", fgColor="2E5FA3")   # blue for metadata cols
_META_FONT      = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
_DATA_FONT      = Font(name="Calibri", size=10)
_ALT_FILL       = PatternFill("solid", fgColor="EAF0FB")
_DEADLINE_FILL  = PatternFill("solid", fgColor="FFF2CC")   # yellow highlight for deadlines
_DEADLINE_FONT  = Font(bold=True, name="Calibri", size=10, color="7B3F00")

# ── Canonical column order ────────────────────────────────────────────────────
# Metadata first, then all data fields, deadlines always last.
_META_COLS = ["source_file", "report_period", "table_type", "table_name"]

_DATA_COLS = [
    "application_id",
    "name_of_applicant",
    "submission_date",
    "region",
    "location_of_project",
    "type_of_project",
    "installed_capacity_mw",
    "first_scod_of_generation_project",
    "connectivity_granted_mw",
    "substation",
    "date_of_connectivity_intimation_in_principle",
    "date_of_connectivity_intimation_final",
    "connectivity_gna_start_date_in_principle",
    "connectivity_gna_start_date_firm",
    # FC-specific
    "criterion_for_applying",
    "revised_criterion",
    "revised_scod",
    "application_status",
]

_DEADLINE_COLS = [
    "due_date_of_fc",
    "due_date_for_submission_of_land_docs",
]

ALL_COLUMNS = _META_COLS + _DATA_COLS + _DEADLINE_COLS

# Human-readable header labels (same order as ALL_COLUMNS)
_HEADER_LABELS = {
    "source_file":                                      "Source File",
    "report_period":                                    "Report Period",
    "table_type":                                       "Table Type",
    "table_name":                                       "Table Name",
    "application_id":                                   "Application ID",
    "name_of_applicant":                                "Name of Applicant",
    "submission_date":                                  "Submission Date",
    "region":                                           "Region",
    "location_of_project":                              "Location of Project",
    "type_of_project":                                  "Type of Project",
    "installed_capacity_mw":                            "Installed Capacity (MW)",
    "first_scod_of_generation_project":                 "First SCOD of Generation Project",
    "connectivity_granted_mw":                          "Connectivity Granted (MW)",
    "substation":                                       "Substation",
    "date_of_connectivity_intimation_in_principle":     "Date of Conn. Intimation (In-principle)",
    "date_of_connectivity_intimation_final":            "Date of Conn. Intimation (Final)",
    "connectivity_gna_start_date_in_principle":         "Connectivity/GNA Start Date (In-principle)",
    "connectivity_gna_start_date_firm":                 "Connectivity/GNA Start Date (Firm)",
    "criterion_for_applying":                           "Criterion for Applying",
    "revised_criterion":                                "Revised Criterion",
    "revised_scod":                                     "Revised SCOD",
    "application_status":                               "Application Status",
    "due_date_of_fc":                                   "Due Date of FC",
    "due_date_for_submission_of_land_docs":             "Due Date for Land Docs",
}


def _auto_size(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col_cells
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 55)


def results_to_excel(
    results: list[FileExtractionResult],
    output_path: Path,
) -> Path:
    """
    Write all extraction results into a single Excel sheet.

    Every row from every table across all files is written to one sheet.
    Metadata columns (source_file, report_period, table_type, table_name)
    are prepended to each row so the sheet is self-contained and filterable.

    Deadline columns are highlighted in amber — they are always last.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CTUIL Compliance Data"

    # ── Header row ────────────────────────────────────────────────────────────
    for col_idx, key in enumerate(ALL_COLUMNS, start=1):
        label = _HEADER_LABELS.get(key, key)
        cell = ws.cell(row=1, column=col_idx, value=label)
        is_meta     = key in _META_COLS
        is_deadline = key in _DEADLINE_COLS
        if is_deadline:
            cell.fill = _DEADLINE_FILL
            cell.font = _DEADLINE_FONT
        elif is_meta:
            cell.fill = _META_FILL
            cell.font = _META_FONT
        else:
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 36

    # ── Data rows ─────────────────────────────────────────────────────────────
    row_idx = 2
    total_written = 0

    for file_result in results:
        for tbl in file_result.tables:
            for row in tbl.rows:
                row_dict = row.model_dump()
                apply_alt = (row_idx % 2 == 0)

                for col_idx, key in enumerate(ALL_COLUMNS, start=1):
                    if key == "source_file":
                        val = file_result.source_file
                    elif key == "table_type":
                        val = tbl.table_type
                    elif key == "table_name":
                        val = tbl.table_name
                    else:
                        val = row_dict.get(key)

                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.font = _DATA_FONT
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    if key in _DEADLINE_COLS and val:
                        cell.fill = PatternFill("solid", fgColor="FFFACD")
                    elif apply_alt:
                        cell.fill = _ALT_FILL

                row_idx += 1
                total_written += 1

    if total_written == 0:
        ws.cell(row=2, column=1, value="No data extracted from the provided PDFs.")
        logger.warning("No rows written to Excel.")

    _auto_size(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions   # enable column filtering

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info("Excel saved: %s  (%d row(s))", output_path, total_written)
    return output_path
